#!/usr/bin/env python3
"""
2026 FIFA World Cup Standings Tracker
Generates index.html for GitHub Pages + Excel locally.
"""

import requests
import json
import os
import time
from datetime import date, datetime, timedelta
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
START_DATE   = date(2026, 6, 11)
DATA_FILE    = os.path.join(SCRIPT_DIR, "wc_games.json")
HTML_FILE    = os.path.join(SCRIPT_DIR, "index.html")
EXCEL_FILE   = os.path.expanduser("~/Downloads/wc_standings.xlsx")
ESPN_URL     = "http://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"
IN_CI        = os.environ.get("GITHUB_ACTIONS") == "true"

FANTASY_TEAMS = {
    "Oliver's Army":                  ["Spain","Portugal","Germany","Mexico","Switzerland","Ecuador","Canada","Sweden","Scotland","South Africa","Bosnia-Herzegovina","Ghana"],
    "Team Chicken Dinner":            ["France","Portugal","Croatia","Mexico","Switzerland","Australia","Norway","Sweden","Scotland","South Africa","Bosnia-Herzegovina","New Zealand"],
    "And It Was All Yellow... Cards": ["France","Brazil","Germany","Mexico","Switzerland","Austria","Norway","Paraguay","Czechia","Qatar","Saudi Arabia","Ghana"],
    "Arbitrage United FC":            ["Argentina","Brazil","Germany","Mexico","Switzerland","Ecuador","Norway","Ivory Coast","Czechia","South Africa","Bosnia-Herzegovina","Ghana"],
    "Andrew Hakes":                   ["England","Brazil","Colombia","Mexico","Iran","South Korea","Canada","Ivory Coast","Congo DR","South Africa","Jordan","Ghana"],
    "MJ's Team":                      ["Spain","Brazil","Germany","Mexico","Switzerland","Ecuador","Canada","Ivory Coast","Scotland","Uzbekistan","Saudi Arabia","Ghana"],
    "The Heeters":                    ["Spain","Netherlands","Germany","Mexico","Switzerland","Ecuador","Canada","Ivory Coast","Scotland","Iraq","Bosnia-Herzegovina","Ghana"],
    "Shai Gorgeous":                  ["Spain","Portugal","Germany","Mexico","Switzerland","South Korea","Canada","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","New Zealand"],
    "Road to Glory":                  ["France","Brazil","Germany","United States","Switzerland","Ecuador","Norway","Sweden","Scotland","South Africa","Bosnia-Herzegovina","Ghana"],
    "Pelley FC 1":                    ["England","Portugal","Germany","Mexico","Türkiye","South Korea","Norway","Ivory Coast","Czechia","South Africa","Saudi Arabia","Ghana"],
    "Aaron Ma":                       ["France","Portugal","Germany","Mexico","Türkiye","South Korea","Norway","Ivory Coast","Czechia","Iraq","Saudi Arabia","Ghana"],
    "Milos Vukovic":                  ["France","Brazil","Germany","Mexico","Switzerland","Ecuador","Norway","Ivory Coast","Tunisia","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "GAM FC":                         ["England","Portugal","Germany","Mexico","Japan","Australia","Canada","Paraguay","Scotland","Uzbekistan","Jordan","Ghana"],
    "Outperformance":                 ["France","Portugal","Germany","Mexico","Switzerland","South Korea","Algeria","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","New Zealand"],
    "I don't know football":          ["Argentina","Portugal","Germany","Mexico","Switzerland","Ecuador","Canada","Paraguay","Congo DR","South Africa","Saudi Arabia","Ghana"],
    "Led_Zep":                        ["Spain","Portugal","Germany","Mexico","Japan","Ecuador","Norway","Sweden","Scotland","South Africa","Bosnia-Herzegovina","Ghana"],
    "Team Matthew":                   ["France","Brazil","Germany","United States","Switzerland","South Korea","Norway","Sweden","Czechia","Uzbekistan","Bosnia-Herzegovina","New Zealand"],
    "Heim's Dimes":                   ["France","Morocco","Germany","Mexico","Japan","South Korea","Canada","Paraguay","Scotland","Qatar","Cape Verde","Curaçao"],
    "Italy should have been here":    ["Spain","Brazil","Germany","Mexico","Switzerland","Ecuador","Norway","Sweden","Czechia","Iraq","Bosnia-Herzegovina","Ghana"],
    "Big Dawgz 2":                    ["France","Portugal","Colombia","Mexico","Japan","Ecuador","Canada","Ivory Coast","Congo DR","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "True Be-Leaf-ers":               ["France","Brazil","Belgium","Uruguay","Japan","South Korea","Canada","Ivory Coast","Scotland","South Africa","Saudi Arabia","Ghana"],
    "Les Rouges 2":                   ["France","Brazil","Colombia","Uruguay","Japan","Austria","Canada","Ivory Coast","Czechia","Qatar","Saudi Arabia","Ghana"],
    "Shai Gorgeous Alexander":        ["England","Portugal","Germany","United States","Japan","South Korea","Canada","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","New Zealand"],
    "Pelley FC 2":                    ["France","Netherlands","Germany","Uruguay","Switzerland","Austria","Egypt","Ivory Coast","Congo DR","Qatar","Bosnia-Herzegovina","Ghana"],
    "Les Rouges 1":                   ["Spain","Portugal","Germany","Senegal","Switzerland","Austria","Canada","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","Ghana"],
    "Canada":                         ["Spain","Portugal","Belgium","United States","Switzerland","Austria","Canada","Ivory Coast","Czechia","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "Big Dawgz":                      ["France","Portugal","Germany","Uruguay","Japan","South Korea","Norway","Sweden","Czechia","Qatar","Saudi Arabia","Ghana"],
    "Saqib Shamasdin":                ["France","Portugal","Belgium","United States","Japan","South Korea","Norway","Ivory Coast","Czechia","Qatar","Saudi Arabia","Ghana"],
    "Dave W's Team":                  ["France","Brazil","Germany","Uruguay","Switzerland","Ecuador","Norway","Sweden","Czechia","Qatar","Bosnia-Herzegovina","Ghana"],
    "GAM Digital":                    ["Spain","Portugal","Germany","United States","Japan","Ecuador","Norway","Sweden","Scotland","South Africa","Bosnia-Herzegovina","Ghana"],
    "Italy":                          ["France","Portugal","Germany","Uruguay","Türkiye","Austria","Norway","Sweden","Scotland","South Africa","Bosnia-Herzegovina","New Zealand"],
    "JHoyle21":                       ["Spain","Portugal","Germany","Uruguay","Switzerland","Ecuador","Norway","Sweden","Scotland","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "Gunners FC":                     ["England","Portugal","Germany","Senegal","Switzerland","Ecuador","Egypt","Sweden","Scotland","Qatar","Bosnia-Herzegovina","Ghana"],
    "E(xtremely) T(alented) F(ootballer)s": ["Spain","Portugal","Colombia","Senegal","Japan","Austria","Canada","Ivory Coast","Scotland","Iraq","Bosnia-Herzegovina","Ghana"],
    "Spain":                          ["Spain","Portugal","Germany","Mexico","Japan","Ecuador","Norway","Paraguay","Scotland","South Africa","Bosnia-Herzegovina","Ghana"],
    "Balham FC":                      ["France","Portugal","Germany","Senegal","Japan","South Korea","Canada","Ivory Coast","Czechia","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "Party of 5":                     ["Spain","Portugal","Germany","Uruguay","Switzerland","South Korea","Norway","Sweden","Czechia","South Africa","Bosnia-Herzegovina","New Zealand"],
    "WhatDoesOffsideEvenMean?":       ["France","Brazil","Germany","Uruguay","Japan","Ecuador","Norway","Ivory Coast","Scotland","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "The Mainoo Event":               ["Spain","Portugal","Belgium","Mexico","Türkiye","Ecuador","Norway","Ivory Coast","Scotland","Iraq","Bosnia-Herzegovina","Ghana"],
    "Team Dreyer":                    ["Spain","Portugal","Germany","Mexico","Türkiye","Ecuador","Norway","Sweden","Czechia","Qatar","Jordan","Ghana"],
    "France":                         ["France","Portugal","Colombia","Uruguay","Japan","South Korea","Canada","Paraguay","Czechia","South Africa","Saudi Arabia","Ghana"],
    "Bryant Spivak":                  ["France","Brazil","Germany","Mexico","Türkiye","Ecuador","Norway","Ivory Coast","Czechia","South Africa","Jordan","New Zealand"],
    "Toronto Inferno":                ["Spain","Netherlands","Germany","Mexico","Japan","Ecuador","Canada","Ivory Coast","Czechia","Uzbekistan","Cape Verde","New Zealand"],
    "Portugal 1966":                  ["Spain","Portugal","Belgium","United States","Türkiye","Austria","Norway","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","Ghana"],
    "Performance":                    ["Spain","Portugal","Belgium","Uruguay","Japan","South Korea","Canada","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","Ghana"],
    "Capital Gains FC":               ["France","Brazil","Germany","Uruguay","Japan","South Korea","Algeria","Ivory Coast","Czechia","Uzbekistan","Saudi Arabia","Ghana"],
    "ROI naldo":                      ["France","Brazil","Germany","Uruguay","Japan","Ecuador","Algeria","Ivory Coast","Scotland","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "Emerging Market Maestro's":      ["Spain","Portugal","Germany","Senegal","Japan","Ecuador","Norway","Ivory Coast","Scotland","Iraq","Saudi Arabia","Ghana"],
    "Competitive Intelligence":       ["Spain","Portugal","Germany","Uruguay","Japan","Ecuador","Norway","Sweden","Czechia","Qatar","Saudi Arabia","Ghana"],
    "Benchmark":                      ["Spain","Netherlands","Germany","Senegal","Iran","South Korea","Norway","Ivory Coast","Czechia","Qatar","Bosnia-Herzegovina","New Zealand"],
    "Berserkers":                     ["Spain","Portugal","Germany","Uruguay","Switzerland","Ecuador","Norway","Sweden","Czechia","Uzbekistan","Saudi Arabia","New Zealand"],
    "Fake It Till You Jake It":       ["Spain","Netherlands","Germany","Senegal","Japan","South Korea","Egypt","Ivory Coast","Czechia","Uzbekistan","Bosnia-Herzegovina","Ghana"],
    "Gov Bonds Intern":               ["Spain","Portugal","Germany","United States","Türkiye","Ecuador","Algeria","Sweden","Czechia","Qatar","Bosnia-Herzegovina","Ghana"],
    "Project 2026":                   ["Spain","Portugal","Germany","Uruguay","Japan","Austria","Norway","Paraguay","Czechia","Uzbekistan","Jordan","Ghana"],
    "Team Chicken Dinner (vibes)":    ["Spain","Morocco","Belgium","Senegal","Japan","South Korea","Canada","Panama","Czechia","Uzbekistan","Cape Verde","Ghana"],
    "Christine Ramsay":               ["Spain","Morocco","Croatia","Mexico","Japan","Ecuador","Algeria","Paraguay","Tunisia","South Africa","Saudi Arabia","Ghana"],
    "Turkey4TheWin":                  ["Spain","Brazil","Germany","Senegal","Türkiye","Ecuador","Norway","Ivory Coast","Tunisia","South Africa","Jordan","New Zealand"],
}


def load_games():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE) as f:
            return json.load(f)
    return {}


def save_games(data):
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)


def fetch_date(d):
    resp = requests.get(ESPN_URL, params={"dates": d.strftime("%Y%m%d")}, timeout=15)
    resp.raise_for_status()
    return resp.json().get("events", [])


def detect_penalties(event, competition):
    status_detail = event.get("status", {}).get("type", {}).get("detail", "").lower()
    if any(k in status_detail for k in ["pk", "pen", "shootout"]):
        return True
    for note in competition.get("notes", []):
        text = " ".join(str(v) for v in note.values()).lower()
        if any(k in text for k in ["pk", "pen", "shootout"]):
            return True
    return False


def parse_event(event):
    status_type = event.get("status", {}).get("type", {})
    if not status_type.get("completed") or status_type.get("state") != "post":
        return None
    comps = event.get("competitions", [])
    if not comps:
        return None
    comp = comps[0]
    competitors = comp.get("competitors", [])
    if len(competitors) < 2:
        return None
    home = next((c for c in competitors if c.get("homeAway") == "home"), competitors[0])
    away = next((c for c in competitors if c.get("homeAway") == "away"), competitors[1])
    try:
        home_score = int(home.get("score", 0))
        away_score = int(away.get("score", 0))
    except (ValueError, TypeError):
        return None
    return {
        "id": event["id"],
        "date": event.get("date", "")[:10],
        "home": home.get("team", {}).get("displayName", "?"),
        "away": away.get("team", {}).get("displayName", "?"),
        "home_score": home_score,
        "away_score": away_score,
        "home_winner": home.get("winner", False),
        "away_winner": away.get("winner", False),
        "penalties": detect_penalties(event, comp),
    }


def build_standings(games):
    teams = {}

    def get(team):
        if team not in teams:
            teams[team] = dict(Team=team, GP=0, W=0, D=0, L=0, GF=0, GA=0, GD=0, Pts=0)
        return teams[team]

    for g in games.values():
        h, a = get(g["home"]), get(g["away"])
        h["GP"] += 1; a["GP"] += 1
        h["GF"] += g["home_score"]; h["GA"] += g["away_score"]
        a["GF"] += g["away_score"]; a["GA"] += g["home_score"]
        if g["penalties"]:
            if g["home_winner"]:
                h["W"] += 1; h["Pts"] += 2; a["L"] += 1; a["Pts"] += 1
            else:
                a["W"] += 1; a["Pts"] += 2; h["L"] += 1; h["Pts"] += 1
        elif g["home_score"] > g["away_score"]:
            h["W"] += 1; h["Pts"] += 3; a["L"] += 1
        elif g["away_score"] > g["home_score"]:
            a["W"] += 1; a["Pts"] += 3; h["L"] += 1
        else:
            h["D"] += 1; h["Pts"] += 1; a["D"] += 1; a["Pts"] += 1

    for t in teams.values():
        t["GD"] = t["GF"] - t["GA"]
    return teams


def build_fantasy_standings(country_standings):
    country_pts = {name: data["Pts"] for name, data in country_standings.items()}
    rows = []
    for team_name, countries in FANTASY_TEAMS.items():
        country_data = [(c, country_pts.get(c, 0)) for c in countries]
        total = sum(pts for _, pts in country_data)
        breakdown = ", ".join(f"{c} ({pts})" for c, pts in country_data)
        rows.append({"Team": team_name, "Total Pts": total, "Countries": breakdown, "CountryData": country_data})
    rows.sort(key=lambda r: r["Total Pts"], reverse=True)
    for i, r in enumerate(rows, 1):
        r["Rank"] = i
    return rows


def write_html(standings, fantasy_rows):
    import json as _json
    now = datetime.utcnow().strftime("%B %d, %Y at %H:%M UTC")

    # Country table rows
    sorted_countries = sorted(standings.values(), key=lambda x: (-x["Pts"], -x["GD"], -x["GF"]))
    country_rows_html = ""
    for i, row in enumerate(sorted_countries, 1):
        gd = row["GD"]
        gd_str = f"+{gd}" if gd > 0 else str(gd)
        gd_cls = "pos" if gd > 0 else ("neg" if gd < 0 else "num")
        country_rows_html += f"""
            <tr>
              <td class="num">{i}</td>
              <td class="name">{row['Team']}</td>
              <td class="num sm">{row['GP']}</td>
              <td class="num">{row['W']}</td>
              <td class="num sm">{row['D']}</td>
              <td class="num sm">{row['L']}</td>
              <td class="num sm">{row['GF']}</td>
              <td class="num sm">{row['GA']}</td>
              <td class="num {gd_cls}">{gd_str}</td>
              <td class="pts">{row['Pts']}</td>
            </tr>"""

    # Fantasy table rows with podium styling
    medals = {1: "&#127945;", 2: "&#129352;", 3: "&#129353;"}
    podium_cls = {1: "gold", 2: "silver", 3: "bronze"}
    fantasy_rows_html = ""
    for r in fantasy_rows:
        rank = r["Rank"]
        medal_html = medals.get(rank, f'<span class="rnum">{rank}</span>')
        row_cls = podium_cls.get(rank, "")
        safe_team = r["Team"].replace("&", "&amp;")
        fantasy_rows_html += f"""
            <tr class="{row_cls}">
              <td class="num medal-cell">{medal_html}</td>
              <td class="name team-link" data-team="{safe_team}">{safe_team}</td>
              <td class="pts">{r['Total Pts']}</td>
            </tr>"""

    # Embed country data for popups
    team_data = {r["Team"]: r["CountryData"] for r in fantasy_rows}
    team_data_json = _json.dumps(team_data, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>2026 FIFA World Cup Standings</title>
  <style>
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ background: #0d1117; color: #e6edf3; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; min-height: 100vh; }}
    header {{ background: linear-gradient(135deg, #0d2137 0%, #1a4a7a 100%); padding: 2.5rem 1.5rem; text-align: center; border-bottom: 2px solid #1f6feb; }}
    header h1 {{ font-size: 2rem; font-weight: 800; letter-spacing: -0.5px; }}
    header p {{ color: #79c0ff; margin-top: 0.4rem; font-size: 0.95rem; }}
    .tabs {{ display: flex; background: #161b22; border-bottom: 1px solid #30363d; padding: 0 1.5rem; }}
    .tab {{ padding: 1rem 1.5rem; cursor: pointer; font-weight: 600; font-size: 0.95rem; color: #8b949e; border-bottom: 3px solid transparent; transition: color 0.15s, border-color 0.15s; user-select: none; white-space: nowrap; }}
    .tab:hover {{ color: #c9d1d9; }}
    .tab.active {{ color: #58a6ff; border-bottom-color: #1f6feb; }}
    .tab-content {{ display: none; padding: 1.5rem; max-width: 900px; margin: 0 auto; }}
    .tab-content.active {{ display: block; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 0.93rem; }}
    thead tr {{ background: #161b22; }}
    th {{ padding: 0.65rem 0.75rem; text-align: center; font-size: 0.72rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; color: #8b949e; border-bottom: 2px solid #21262d; white-space: nowrap; }}
    th.left {{ text-align: left; }}
    td {{ padding: 0.7rem 0.75rem; border-bottom: 1px solid #21262d; }}
    tbody tr:hover td {{ background: #1c2128; }}
    tbody tr:last-child td {{ border-bottom: none; }}
    .num {{ text-align: center; color: #c9d1d9; }}
    .name {{ text-align: left; font-weight: 500; }}
    .pts {{ text-align: center; font-weight: 700; color: #58a6ff; }}
    .pos {{ color: #3fb950; font-weight: 600; }}
    .neg {{ color: #f85149; font-weight: 600; }}
    .updated {{ text-align: center; color: #484f58; font-size: 0.78rem; padding: 2rem 1rem; border-top: 1px solid #21262d; margin-top: 1rem; }}
    .rnum {{ color: #8b949e; font-size: 0.85rem; }}

    /* Podium rows */
    tr.gold td {{ background: linear-gradient(90deg, rgba(255,215,0,0.1) 0%, rgba(255,215,0,0.02) 60%, transparent 100%); }}
    tr.silver td {{ background: linear-gradient(90deg, rgba(192,192,192,0.08) 0%, rgba(192,192,192,0.01) 60%, transparent 100%); }}
    tr.bronze td {{ background: linear-gradient(90deg, rgba(205,127,50,0.1) 0%, rgba(205,127,50,0.02) 60%, transparent 100%); }}
    tr.gold td:first-child {{ border-left: 3px solid #ffd700; }}
    tr.silver td:first-child {{ border-left: 3px solid #c0c0c0; }}
    tr.bronze td:first-child {{ border-left: 3px solid #cd7f32; }}
    tr.gold:hover td, tr.silver:hover td, tr.bronze:hover td {{ filter: brightness(1.2); }}
    .medal-cell {{ font-size: 1.2rem; text-align: center; }}
    .team-link {{ cursor: pointer; color: #e6edf3 !important; }}
    .team-link:hover {{ color: #58a6ff !important; text-decoration: underline; }}
    tr.gold .team-link {{ color: #ffd700 !important; font-weight: 700; }}
    tr.silver .team-link {{ color: #c0c0c0 !important; font-weight: 700; }}
    tr.bronze .team-link {{ color: #cd7f32 !important; font-weight: 700; }}
    tr.gold:hover .team-link, tr.silver:hover .team-link, tr.bronze:hover .team-link {{ text-decoration: underline; }}

    /* Modal */
    .overlay {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.75); z-index: 100; align-items: center; justify-content: center; backdrop-filter: blur(3px); }}
    .overlay.open {{ display: flex; }}
    .modal {{ background: #161b22; border: 1px solid #30363d; border-radius: 14px; padding: 1.5rem; width: 92%; max-width: 420px; max-height: 88vh; overflow-y: auto; animation: pop 0.18s ease; }}
    @keyframes pop {{ from {{ opacity:0; transform:scale(0.93) translateY(8px); }} to {{ opacity:1; transform:scale(1) translateY(0); }} }}
    .modal-head {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 1.2rem; }}
    .modal-title {{ font-size: 1.05rem; font-weight: 700; line-height: 1.3; padding-right: 1rem; }}
    .modal-close {{ background: #21262d; border: none; color: #8b949e; width: 28px; height: 28px; border-radius: 50%; font-size: 1rem; cursor: pointer; flex-shrink: 0; display: flex; align-items: center; justify-content: center; }}
    .modal-close:hover {{ background: #30363d; color: #e6edf3; }}
    .country-row {{ display: flex; justify-content: space-between; align-items: center; padding: 0.55rem 0; border-bottom: 1px solid #21262d; gap: 0.5rem; }}
    .country-row:last-of-type {{ border-bottom: none; }}
    .c-name {{ font-size: 0.9rem; color: #c9d1d9; }}
    .c-pts {{ font-size: 0.85rem; font-weight: 700; padding: 0.15rem 0.5rem; border-radius: 5px; min-width: 40px; text-align: center; }}
    .c-pts.hi {{ color: #3fb950; background: rgba(63,185,80,0.12); }}
    .c-pts.mid {{ color: #d29922; background: rgba(210,153,34,0.12); }}
    .c-pts.lo {{ color: #6e7681; background: rgba(110,118,129,0.1); }}
    .modal-total {{ display: flex; justify-content: space-between; align-items: center; padding-top: 0.9rem; margin-top: 0.5rem; border-top: 2px solid #30363d; font-weight: 700; font-size: 1rem; }}
    .modal-total span:last-child {{ color: #58a6ff; font-size: 1.2rem; }}

    /* Compare tab */
    .cmp-bar {{ display: flex; align-items: center; gap: 0.75rem; margin-bottom: 1.75rem; flex-wrap: wrap; }}
    .team-select {{ flex: 1; min-width: 180px; background: #161b22; color: #e6edf3; border: 1px solid #30363d; border-radius: 8px; padding: 0.65rem 0.8rem; font-size: 0.9rem; cursor: pointer; appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 12 12'%3E%3Cpath fill='%238b949e' d='M6 8L1 3h10z'/%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 0.75rem center; padding-right: 2rem; }}
    .team-select:focus {{ outline: none; border-color: #58a6ff; box-shadow: 0 0 0 3px rgba(31,111,235,0.15); }}
    .vs-badge {{ font-weight: 900; color: #484f58; font-size: 0.85rem; letter-spacing: 0.05em; flex-shrink: 0; }}
    .cmp-empty {{ text-align: center; color: #484f58; padding: 3.5rem 1rem; font-size: 0.95rem; border: 1px dashed #21262d; border-radius: 10px; }}
    .cmp-heads {{ display: grid; grid-template-columns: 1fr 160px 1fr; align-items: center; padding: 0.6rem 0.5rem 0.9rem; border-bottom: 2px solid #30363d; gap: 0.5rem; }}
    .cmp-hname {{ font-weight: 700; font-size: 0.9rem; color: #e6edf3; line-height: 1.3; }}
    .cmp-hname.r {{ text-align: right; }}
    .cmp-hmid {{ text-align: center; font-size: 0.68rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.07em; color: #484f58; }}
    .cmp-row {{ display: grid; grid-template-columns: 1fr 160px 1fr; align-items: center; padding: 0.5rem 0.5rem; border-bottom: 1px solid #1a1f26; gap: 0.5rem; border-radius: 4px; }}
    .cmp-row.shared {{ background: rgba(63,185,80,0.05); border-left: 2px solid rgba(63,185,80,0.25); }}
    .cmp-side {{ display: flex; }}
    .cmp-side.r {{ justify-content: flex-end; }}
    .cmp-country {{ text-align: center; font-size: 0.83rem; color: #8b949e; padding: 0 0.25rem; }}
    .cmp-row.shared .cmp-country {{ color: #c9d1d9; }}
    .ptag {{ font-size: 0.82rem; font-weight: 700; padding: 0.18rem 0.5rem; border-radius: 5px; }}
    .ptag.hi {{ color: #3fb950; background: rgba(63,185,80,0.12); }}
    .ptag.mid {{ color: #d29922; background: rgba(210,153,34,0.12); }}
    .ptag.lo {{ color: #6e7681; background: rgba(110,118,129,0.1); }}
    .ptag.none {{ color: #30363d; }}
    .cmp-totals {{ display: grid; grid-template-columns: 1fr 160px 1fr; padding: 0.9rem 0.5rem 0.5rem; border-top: 2px solid #30363d; margin-top: 0.25rem; gap: 0.5rem; align-items: center; }}
    .cmp-tot {{ font-weight: 800; font-size: 1.1rem; color: #58a6ff; }}
    .cmp-tot.r {{ text-align: right; }}
    .cmp-tot-mid {{ text-align: center; font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.07em; color: #484f58; font-weight: 600; }}
    .cmp-legend {{ display: flex; gap: 1rem; margin-top: 1.25rem; flex-wrap: wrap; }}
    .legend-item {{ display: flex; align-items: center; gap: 0.4rem; font-size: 0.75rem; color: #6e7681; }}
    .legend-dot {{ width: 8px; height: 8px; border-radius: 50%; }}

    @media (max-width: 640px) {{
      .sm {{ display: none; }}
      header h1 {{ font-size: 1.4rem; }}
      .tab {{ padding: 0.75rem 0.7rem; font-size: 0.8rem; }}
      .tab-content {{ padding: 0.75rem 0.25rem; }}
      td, th {{ padding: 0.55rem 0.4rem; }}
      .cmp-heads, .cmp-row, .cmp-totals {{ grid-template-columns: 1fr 120px 1fr; }}
      .cmp-country {{ font-size: 0.75rem; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>&#9917; 2026 FIFA World Cup</h1>
    <p>Live Standings &mdash; Custom Points System</p>
  </header>

  <div class="tabs">
    <div class="tab active" onclick="switchTab('country', this)">&#127758; Country Standings</div>
    <div class="tab" onclick="switchTab('fantasy', this)">&#127942; Standings</div>
    <div class="tab" onclick="switchTab('compare', this)">&#9889; Compare</div>
  </div>

  <div id="country" class="tab-content active">
    <table>
      <thead>
        <tr>
          <th>#</th><th class="left">Country</th>
          <th class="sm">GP</th><th>W</th><th class="sm">D</th><th class="sm">L</th>
          <th class="sm">GF</th><th class="sm">GA</th><th>GD</th><th>Pts</th>
        </tr>
      </thead>
      <tbody>{country_rows_html}
      </tbody>
    </table>
  </div>

  <div id="fantasy" class="tab-content">
    <table>
      <thead>
        <tr>
          <th>#</th><th class="left">Team</th><th>Pts</th>
        </tr>
      </thead>
      <tbody>{fantasy_rows_html}
      </tbody>
    </table>
  </div>

  <div id="compare" class="tab-content">
    <div class="cmp-bar">
      <select id="sel1" class="team-select"><option value="">Select a team...</option></select>
      <div class="vs-badge">VS</div>
      <select id="sel2" class="team-select"><option value="">Select a team...</option></select>
    </div>
    <div id="cmp-placeholder" class="cmp-empty">&#128101; Select two teams above to compare their picks</div>
    <div id="cmp-result" style="display:none">
      <div class="cmp-heads">
        <div id="h1" class="cmp-hname"></div>
        <div class="cmp-hmid">Country</div>
        <div id="h2" class="cmp-hname r"></div>
      </div>
      <div id="cmp-rows"></div>
      <div class="cmp-totals">
        <div id="tot1" class="cmp-tot"></div>
        <div class="cmp-tot-mid">Total Pts</div>
        <div id="tot2" class="cmp-tot r"></div>
      </div>
      <div class="cmp-legend">
        <div class="legend-item"><div class="legend-dot" style="background:rgba(63,185,80,0.4)"></div> Both teams have this country</div>
        <div class="legend-item"><div class="legend-dot" style="background:#30363d"></div> Only one team has this country</div>
      </div>
    </div>
  </div>

  <p class="updated">Last updated: {now} &nbsp;&bull;&nbsp; Win=3pts &nbsp;&bull;&nbsp; Draw=1pt &nbsp;&bull;&nbsp; Penalties: W=2pts / L=1pt &nbsp;&bull;&nbsp; Click a team name to see their picks</p>

  <!-- Modal -->
  <div class="overlay" id="overlay" onclick="handleOverlayClick(event)">
    <div class="modal" id="modal">
      <div class="modal-head">
        <div class="modal-title" id="modal-title"></div>
        <button class="modal-close" onclick="closeModal()">&#x2715;</button>
      </div>
      <div id="modal-body"></div>
      <div class="modal-total">
        <span>Total Points</span>
        <span id="modal-total"></span>
      </div>
    </div>
  </div>

  <script>
    const TEAM_DATA = {team_data_json};

    function switchTab(name, el) {{
      document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
      document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
      document.getElementById(name).classList.add('active');
      el.classList.add('active');
    }}

    // ── Standings modal ──────────────────────────────────────────
    function showTeam(teamName) {{
      const data = TEAM_DATA[teamName];
      if (!data) return;
      document.getElementById('modal-title').textContent = teamName;
      const sorted = [...data].sort((a, b) => b[1] - a[1]);
      let total = 0, html = '';
      sorted.forEach(([country, pts]) => {{
        total += pts;
        const cls = pts >= 3 ? 'hi' : pts >= 1 ? 'mid' : 'lo';
        html += `<div class="country-row"><span class="c-name">${{country}}</span><span class="c-pts ${{cls}}">${{pts}} pts</span></div>`;
      }});
      document.getElementById('modal-body').innerHTML = html;
      document.getElementById('modal-total').textContent = total + ' pts';
      document.getElementById('overlay').classList.add('open');
    }}
    function closeModal() {{ document.getElementById('overlay').classList.remove('open'); }}
    function handleOverlayClick(e) {{ if (e.target === document.getElementById('overlay')) closeModal(); }}
    document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closeModal(); }});
    document.querySelectorAll('.team-link').forEach(el => {{
      el.addEventListener('click', () => showTeam(el.dataset.team));
    }});

    // ── Compare tab ──────────────────────────────────────────────
    const sortedTeams = Object.keys(TEAM_DATA).sort((a, b) => a.localeCompare(b));
    ['sel1','sel2'].forEach(id => {{
      const el = document.getElementById(id);
      sortedTeams.forEach(name => {{
        const o = document.createElement('option');
        o.value = name; o.textContent = name;
        el.appendChild(o);
      }});
      el.addEventListener('change', runCompare);
    }});

    function ptag(pts) {{
      if (pts === undefined) return '<span class="ptag none">—</span>';
      const cls = pts >= 3 ? 'hi' : pts >= 1 ? 'mid' : 'lo';
      return `<span class="ptag ${{cls}}">${{pts}} pts</span>`;
    }}

    function runCompare() {{
      const t1 = document.getElementById('sel1').value;
      const t2 = document.getElementById('sel2').value;
      if (!t1 || !t2) return;

      const d1 = Object.fromEntries(TEAM_DATA[t1]);
      const d2 = Object.fromEntries(TEAM_DATA[t2]);
      const all = [...new Set([...Object.keys(d1), ...Object.keys(d2)])];

      // Sort: shared first (by pts desc), then exclusives
      all.sort((a, b) => {{
        const sa = (a in d1) && (a in d2);
        const sb = (b in d1) && (b in d2);
        if (sa !== sb) return sa ? -1 : 1;
        return Math.max(d2[b]||0, d1[b]||0) - Math.max(d2[a]||0, d1[a]||0);
      }});

      let t1Tot = 0, t2Tot = 0, html = '';
      all.forEach(country => {{
        const p1 = d1[country], p2 = d2[country];
        const shared = p1 !== undefined && p2 !== undefined;
        if (p1 !== undefined) t1Tot += p1;
        if (p2 !== undefined) t2Tot += p2;
        html += `<div class="cmp-row${{shared ? ' shared' : ''}}">
          <div class="cmp-side">${{ptag(p1)}}</div>
          <div class="cmp-country">${{country}}</div>
          <div class="cmp-side r">${{ptag(p2)}}</div>
        </div>`;
      }});

      document.getElementById('h1').textContent = t1;
      document.getElementById('h2').textContent = t2;
      document.getElementById('cmp-rows').innerHTML = html;
      document.getElementById('tot1').textContent = t1Tot + ' pts';
      document.getElementById('tot2').textContent = t2Tot + ' pts';
      document.getElementById('cmp-result').style.display = 'block';
      document.getElementById('cmp-placeholder').style.display = 'none';
    }}
  </script>
</body>
</html>"""

    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  Saved -> index.html")


def write_excel(standings, fantasy_rows):
    cols = ["Team", "GP", "W", "D", "L", "GF", "GA", "GD", "Pts"]
    df = pd.DataFrame(list(standings.values()), columns=cols)
    df = df.sort_values(["Pts", "GD", "GF"], ascending=False).reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df) + 1))
    df2 = pd.DataFrame(fantasy_rows, columns=["Rank", "Team", "Total Pts", "Countries"])

    def style_sheet(ws, team_col=1):
        hf = PatternFill("solid", fgColor="1F4E79")
        hfont = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = hf; cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = PatternFill("solid", fgColor="D9E1F2" if row_idx % 2 == 0 else "FFFFFF")
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[team_col].alignment = Alignment(horizontal="left", vertical="center")
            row[-1].font = Font(bold=True)
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 4, 8), 60)
        ws.freeze_panes = "A2"

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Country Standings", index=False)
        df2.to_excel(writer, sheet_name="Fantasy Standings", index=False)
        style_sheet(writer.sheets["Country Standings"])
        style_sheet(writer.sheets["Fantasy Standings"])
        ws2 = writer.sheets["Fantasy Standings"]
        for row in ws2.iter_rows(min_row=2):
            row[3].alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["D"].width = 80
        for i in range(2, ws2.max_row + 1):
            ws2.row_dimensions[i].height = 40
    print(f"  Saved -> {EXCEL_FILE}")


def main():
    games = load_games()
    new_count = 0
    today = date.today()
    d = START_DATE

    print("2026 FIFA World Cup Tracker")
    print(f"Scanning {START_DATE} -> {today} for finalized games...\n")

    while d <= today:
        try:
            events = fetch_date(d)
            for event in events:
                gid = event.get("id")
                if gid in games:
                    continue
                parsed = parse_event(event)
                if parsed:
                    games[parsed["id"]] = parsed
                    new_count += 1
                    pk = " [PKs]" if parsed["penalties"] else ""
                    print(f"  + {parsed['date']}  {parsed['home']} {parsed['home_score']}-{parsed['away_score']} {parsed['away']}{pk}")
        except requests.RequestException as e:
            print(f"  [{d}] Request error: {e}")
        time.sleep(0.3)
        d += timedelta(days=1)

    save_games(games)
    if new_count:
        print(f"\n  {new_count} new game(s) added.")
    else:
        print("  No new games since last run.")
    print(f"  Total games tracked: {len(games)}\n")

    standings = build_standings(games)
    if not standings:
        print("No completed games found yet.")
        return

    fantasy_rows = build_fantasy_standings(standings)

    write_html(standings, fantasy_rows)

    if not IN_CI:
        write_excel(standings, fantasy_rows)

    print("\n  === FANTASY STANDINGS ===\n")
    print(f"  {'Rank':<5} {'Team':<42} {'Pts':>5}")
    print(f"  {'-'*54}")
    for r in fantasy_rows:
        print(f"  {r['Rank']:<5} {r['Team']:<42} {r['Total Pts']:>5}")
    print()


if __name__ == "__main__":
    main()
